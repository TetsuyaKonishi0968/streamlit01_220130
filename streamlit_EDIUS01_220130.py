#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st

import subprocess
import openpyxl as excel
import os

import time,subprocess,platform
import pyautogui as pa
import pyperclip

import pandas as pd

import tkinter as tk
import tkinter.messagebox as mb
import tkinter.filedialog as fd

import xlwings as xw
import tempfile




st.title('EDIUS 荒編集アプリ')
"""
## スライド画像の仮配置を自動実行します
EDIUSの右上を露出させてください  
仮配置タイムシート、スライド画像保存フォルダを選択してください

"""
#タイムコード指示エクセルファイルの取得→ 始動


uploaded_file = st.file_uploader("choose an excel file...", type='xlsx')
if uploaded_file is not None:
    
#      bookO = excel.load_workbook(uploaded_file, data_only=True)
#      sheetO = bookO.active
    
#     #元の数式
#     sheetO["D2"] = 'デュレーション'
#     sheetO["D3"] = '=LEFT(TEXT(ABS(LEFT(C3,8)-LEFT(B3,8)+(RIGHT(C3,2)-RIGHT(B3,2))/30*"0:0:1"),"hh:mm:ss.00"),8) & TEXT(INT(RIGHT(TEXT(ABS(LEFT(C3,8)-LEFT(B3,8)+(RIGHT(C3,2)-RIGHT(B3,2))/30*"0:0:1"),"hh:mm:ss.00"),3)*30),"!:00")' 

#     #オートフィル
#     for row_no in range(4, sheetO.max_row + 1):
#         cell_no = f'D{row_no}'
#         sheetO[cell_no] = Translator(sheetO['D3'].value,origin='D3').translate_formula(cell_no)


#     st.markdown("ok")
#     cwd = os.getcwd()

#     st.markdown(cwd)

#     p = os.path.dirname(os.path.abspath("__file__"))
#     st.markdown(p)    
        
#     bookO.save("susiki.xlsx")

#     path = "susiki.xlsx"
#     wb = xw.Book(path)
#     wb.save()  # ブックの保存

#     app = xw.apps.active # ファイルを閉じる
#     xl = xw.apps.active.api
#     xl.Quit()


    book = excel.load_workbook(uploaded_file, data_only=True)
    sheet = book.active

    lis = [cell.value for cell in sheet["A:A"] if cell.value is not None]

    #print(lis)
    #print(len(lis))
    #a = len(lis)
    #print(lis[a-1])

    b = book['Sheet1'].max_row
    #print(b)

    #リストを取得
    it = sheet.iter_rows(
            min_row=3, min_col=1,
            max_row= b, max_col=4)

    eList = []

    for row in it:
        r = []

        for cell in row:
            r.append(cell.value)

        #print(r)
        if r[1] is not None:
            eList.append(r)

    st.markdown(eList)
    st.markdown(len(eList))
    
    #st.markdown(dirpath)

    #dpfn = dirpath +'/'+ eList[0][0]+'.png'
    #dpfn.replace('/', '\\')
    #print(dpfn)

    os.chdir(p)

    st.markdown(os.getcwd())


    #スライド画像フォルダの取得

    # topmost指定(最前面)
    root = tk.Tk()

    root.attributes('-topmost', True)
    root.withdraw()
    root.lift()
    root.focus_force()

    dirpath = fd.askdirectory(
        title='スライド画像フォルダを指定して下さい',initialdir='./')

    #mb.showinfo('対象フォルダ',dirpath)

    st.markdown(dirpath)


    for j in range(len(eList)):    
        pos1 = None
        for i in range(10):
            pos1 = pa.locateOnScreen('edius_parts001.png',grayscale=True, confidence=0.9)
            if pos1 is None:
                time.sleep(1)
                print('探しています')
                continue
            break
        if pos1 is None:
            pa.alert('見つかりません')
            pass
        print('見つかりました:', pos1)

        x1, y1 = pa.center(pos1)
        pa.click(x1, y1)
        time.sleep(1)

        #パス付きファイル名を取得
        #print(dirpath +'/'+ eList[j][0]+'.png')

        dpfn = dirpath +'/'+ eList[j][0]+'.png'
        dpbfn = dpfn.replace('/', '\\')
        print(dpbfn)

        #クリップボードに保存からの貼り付け
        pyperclip.copy(dpbfn)
        pa.hotkey('ctrl', 'v')

        pa.press('enter')

        time.sleep(1)

        pa.hotkey('ctrl', 'c')


        #タイムコード

        print(pa.position())

        pos5 = None
        for i in range(10):
            pos5 = pa.locateOnScreen('edius_parts005.png',grayscale=True, confidence=0.9)
            if pos5 is None:
                time.sleep(1)
                print('探しています')
                continue
            break
        if pos5 is None:
            pa.alert('見つかりません')
            pass
        print('見つかりました:', pos5)

        x5, y5 = pa.center(pos5)

        pyperclip.copy(eList[j][1])

        pa.rightClick(x5+60, y5)
        pa.press('p')

        pos6 = None
        for i in range(10):
            pos6 = pa.locateOnScreen('edius_parts006.png',grayscale=True, confidence=0.9)
            if pos6 is None:
                time.sleep(1)
                print('探しています')
                continue
            break
        if pos6 is None:
            pa.alert('見つかりません')
            pass
        print('見つかりました:', pos6)

        x6, y6 = pa.center(pos6)
        pa.click(x6+100, y6)
        pa.click(x6+100, y6)


        time.sleep(1)
        pa.hotkey('ctrl', 'v')

        time.sleep(1)
        pa.hotkey('alt', 'u')

        pyperclip.copy(eList[j][3])
        pa.hotkey('ctrl', 'v')

        time.sleep(1)

        pa.press('enter')


    time.sleep(1)
    pa.hotkey('ctrl', 's')

    os.chdir(p)
    os.remove("susiki.xlsx")

    pa.alert('完了しました！')

